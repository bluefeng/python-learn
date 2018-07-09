from openpyxl import Workbook, load_workbook
import sys
def try_to_number(s):
	try:
		return int(s)
	except ValueError:
		try:
			return float(s)
		except ValueError:
			pass
	return s

def printData(data):
	for temp in data:
		print(*temp, sep="\t")
def dictToList(dictData):
	result = []
	for k , v in dictData.items() :
		if isinstance(v, list):
			result.append([k] + v)
		else:
			result.append([k , v])
	return result
def saveToSheet(wb, name, data):
	if name == 'Sheet' :
		name = name + len(data)
		print("sheet name cant is Sheet")
	sheet = None
	if 'Sheet' in wb.sheetnames:
		sheet = wb["Sheet"]
		sheet.title = name
	else:
		sheet = wb.create_sheet(name)

	for i in range(len(data)):
		temp = data[i]
		for j in range(len(temp)):
			sheet.cell(row = i + 1, column = j + 1, value = try_to_number(temp[j]))

def saveToXlsx(path, data):
	wb = Workbook()
	saveToSheet(wb, "sheet", data)
	wb.save(path + ".xlsx")

def saveLineToCsv(f, line):
	temp = []
	for one in line:
		temp.append(str(one))
	f.write(",".join(temp) + "\n")

def saveListToCsv(f, data):
	for temp in data:
		saveLineToCsv(f, temp)

def saveToCsv(path, data):
	with open(path + ".csv", 'w') as f:
		saveListToCsv(f, data)

def depthPmsToList(dictPms):
	result = []
	if isinstance(dictPms, list) :
		for temp in dictPms:
			result += depthPmsToList(temp)
	elif isinstance(dictPms, dict):
		for k, v in dictPms.items():
			for temp in depthPmsToList(v) :
				result.append([k, *temp])
	else:
		result.append([dictPms])
	return result

def getListByData(data, path = [], *values):
	tempData = data
	for index in range(len(path)):
		tempData = tempData[path[index]]
	if len(values) == 0 :
		return tempData
	else:
		result = []
		for temp in tempData :
			oneData = []
			for value in values :
				temp2 = temp
				if isinstance(value, list) :
					for index in range(len(value)):
						temp2 = temp2[value[index]]
					oneData.append(temp2)
				else:
					oneData.append(temp2[value])
			result.append(oneData)
		return result

def getDictByData(data, path = [], key = None, *values):
	tempData = data
	for index in range(len(path)):
		tempData = tempData[path[index]]
	if not key or len(values) == 0 :
		return tempData
	else:
		result = {}
		for temp in tempData :
			oneData = []
			curKey = temp
			if isinstance(key, list) :
				for index in range(len(key)):
					curKey = curKey[key[index]]
			else:
				curKey = curKey[key]
			for value in values :
				temp2 = temp
				if isinstance(value, list) :
					for index in range(len(value)):
						temp2 = temp2[value[index]]
					oneData.append(temp2)
				else:
					oneData.append(temp2[value])
			if len(values) == 1 :
				oneData = oneData[0]
			result[curKey] = oneData
		return result

#统计data 字典 中 value 的分布
def getValueCount(dictData, needPrint = None):
	result = {}
	for k , v in dictData.items():
		if not v in result :
			result[v] = 1 
		else:
			result[v] += 1
	if needPrint :
		return dictToList(result)
	else:
		return result

def addValue(add, addTo):
	if type(add) != type(addTo) :
		raise
	if isinstance(add, list):
		for index in range(len(add)) :
			if index >= len(addTo):
				addTo.extend(add[index:len(add)])
				break
			else:
				addTo[index] = addValue(add[index], addTo[index])

	elif isinstance(add, dict):
		for k, v in add.items():
			if not k in addTo :
				addTo[k] = v
			else:
				addTo[k] = addValue(v, addTo[k])
	else:
		addTo += add
	return addTo


def dictAddDictValue(addDict, preDict = {}):
	for k, v in addDict.items():
		if not k in preDict :
			preDict[k] = v
		else:
			preDict[k] += v
	return preDict

def getInputTime():
	time = []
	try :
		time.append(sys.argv[1])
	except:
		pass
	try : 
		time.append(sys.argv[2])
	except:
		pass

	if len(time) == 0 :
		return None
	elif len(time) == 1:
		return time[0]
	else:
		return time
